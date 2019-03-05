import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

with open('/home/adamluo/code/cool/cargo/out_HK.json') as f:
    contents = json.loads(f.read())
lines = {'2015':list(), '2016':list(), '2017':list(), '2018':list()}
for x in contents:
    lines[x['Year']].append(x)
wb = Workbook()
filename = 'newHK.xlsx'

for year in lines.keys():
    ws = wb.create_sheet(title=year)

    for line in tqdm(lines[year]):
        r = [
            line['Corp'],
            line['Shipper_Name'],
            line['Shipper_City_State'] + ', ' + line['Shipper_Country'],
            line['Shipper_Address'] + ' ' + line['Shipper_Address2'],
            line['Shipper_Email'],
            line['Shipper_Phone'],
            line['Consignee_Name'],
            line['Consignee_Address'] + ' ' + line['Consignee_Address2'],
            line['Consignee_Email'],
            line['Consignee_Phone'],
        ]
        count_20, count_40 = 0, 0
        des = ''
        for x in line['Freight']:
            if x['Type'][:2] is '20':
                count_20 += 1
            else:
                count_40 += 1
        des += len(line['Freight']) > 0 and line['Freight'][0]['Description'] or ''
        r.append(str(count_20))
        r.append(str(count_40))
        r.append(str(len(line['Vehicles'])))
        des += len(line['Vehicles']) > 0 and line['Vehicles'][0]['Brand'] or ''
        r.append(len(line['Bulk']) > 0 and (line['Bulk']['Weight'] + '/' + line['Bulk']['Vol']) or '0/0')
        des += len(line['Bulk']) > 0 and line['Bulk']['Description'] or ''
        r.append(des)

        ws.append(r)

wb.save(filename=filename)






